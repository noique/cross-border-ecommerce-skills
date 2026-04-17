# data_processing.py
import pandas as pd
import os
import shutil
from utils import create_directory

def save_to_csv(df, filename):
    """保存数据到CSV文件"""
    original_filename = filename
    file_base, file_ext = os.path.splitext(filename)
    attempt = 0
    while attempt < 10:
        try:
            df.to_csv(filename, index=False, encoding='utf-8-sig')
            print(f"Saved {len(df)} reviews to {filename}")
            return filename
        except PermissionError:
            attempt += 1
            filename = f"{file_base}-{attempt}{file_ext}"
            print(f"File access error. Trying: {filename}")
        except Exception as e:
            attempt += 1
            filename = f"{file_base}-{attempt}{file_ext}"
            print(f"Error saving file: {e}. Trying: {filename}")
    print(f"Failed to save after {attempt} attempts.")
    return None

def format_excel_file(file_path):
    """格式化Excel文件"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment
        from openpyxl.utils import get_column_letter
        
        df = pd.read_csv(file_path)
        original_count = len(df)
        df = df.dropna(subset=['review']).reset_index(drop=True)
        removed_count = original_count - len(df)
        print(f"Removed {removed_count} rows with empty reviews")
        
        df.columns = [col.capitalize() for col in df.columns]
        excel_path = file_path.replace('.csv', '.xlsx')
        df.to_excel(excel_path, index=False)
        
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        
        for cell in ws[1]:
            cell.font = Font(bold=True)
        
        for i in range(1, 5):
            ws.column_dimensions[get_column_letter(i)].width = 15
        ws.column_dimensions['E'].width = 150
        
        for row in ws.rows:
            ws.row_dimensions[row[0].row].height = 35
            for i, cell in enumerate(row, 1):
                if i <= 4:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                else:
                    cell.alignment = Alignment(vertical='center')
        
        wb.save(excel_path)
        print(f"Formatted Excel file saved to: {excel_path}")
        
        txt_path = file_path.replace('.csv', '.txt')
        with open(txt_path, 'w', encoding='utf-8') as txt_file:
            txt_file.write('\t'.join(df.columns) + '\n')
            for _, row in df.iterrows():
                row_values = [str(val) if pd.notna(val) else '' for val in row.values]
                txt_file.write('\t'.join(row_values) + '\n')
        print(f"Exported text file to: {txt_path}")
        
        os.remove(file_path)
        print(f"Removed original CSV file: {file_path}")
        
        return excel_path, df
    except Exception as e:
        print(f"Error formatting Excel file: {e}")
        return file_path, None

def filter_valid_countries(df, country_col):
    """过滤有效国家代码"""
    from config import VALID_COUNTRY_CODES
    
    valid_countries_df = df.copy()
    valid_countries_df[country_col] = valid_countries_df[country_col].str.upper()
    valid_countries_mask = valid_countries_df[country_col].isin(VALID_COUNTRY_CODES)
    valid_countries_df = valid_countries_df[valid_countries_mask]
    
    filtered_count = len(df) - len(valid_countries_df)
    print(f"Filtered out {filtered_count} rows with invalid country codes")
    
    return valid_countries_df

def organize_analysis_files(run_dir):
    """重新组织分析文件"""
    analysis_dir = os.path.join(run_dir, "analysis_results")
    create_directory(analysis_dir)
    
    rating_dir = os.path.join(analysis_dir, "rating_analysis")
    sentiment_dir = os.path.join(analysis_dir, "sentiment_analysis")
    topic_dir = os.path.join(analysis_dir, "topic_analysis")
    country_dir = os.path.join(analysis_dir, "country_analysis")
    time_dir = os.path.join(analysis_dir, "time_analysis")
    word_dir = os.path.join(analysis_dir, "word_analysis")
    
    for directory in [rating_dir, sentiment_dir, topic_dir, country_dir, time_dir, word_dir]:
        create_directory(directory)
    
    for root, dirs, files in os.walk(run_dir):
        if root == run_dir or root == analysis_dir:
            continue
        for file in files:
            file_path = os.path.join(root, file)
            file_lower = file.lower()
            
            # 首先处理评分相关的词云文件
            if ('rating' in file_lower and ('word' in file_lower or 'cloud' in file_lower)):
                shutil.move(file_path, os.path.join(word_dir, file))
            # 然后处理其他词云和词频文件
            elif ('word' in file_lower or 'cloud' in file_lower or 'word_sentiment' in file_lower):
                shutil.move(file_path, os.path.join(word_dir, file))
            # 处理评分相关文件（不包含词云和主题）
            elif 'rating' in file_lower and 'topic' not in file_lower:
                shutil.move(file_path, os.path.join(rating_dir, file))
            # 处理情感分析文件
            elif any(pattern in file_lower for pattern in ['sentiment', 'length_sentiment', 'with_sentiment']):
                shutil.move(file_path, os.path.join(sentiment_dir, file))
            # 处理主题分析文件
            elif 'topic' in file_lower:
                shutil.move(file_path, os.path.join(topic_dir, file))
            # 处理国家分析文件
            elif 'country' in file_lower:
                shutil.move(file_path, os.path.join(country_dir, file))
            # 处理时间趋势文件
            elif 'time' in file_lower or 'trend' in file_lower:
                shutil.move(file_path, os.path.join(time_dir, file))
            # 处理HTML文件
            elif file_lower.endswith('.html'):
                if 'country' in file_lower:
                    shutil.move(file_path, os.path.join(country_dir, file))
                elif 'report' in file_lower:
                    shutil.move(file_path, os.path.join(analysis_dir, file))
    
    # 删除原始分析目录
    original_dirs = ['rating_analysis', 'sentiment_analysis', 'topic_analysis', 
                    'country_analysis', 'time_analysis', 'wordcloud_analysis']
    for dir_name in original_dirs:
        dir_path = os.path.join(run_dir, dir_name)
        try:
            if os.path.exists(dir_path) and os.path.isdir(dir_path):
                shutil.rmtree(dir_path)
        except Exception as e:
            print(f"Error removing directory {dir_path}: {e}")
    
    print(f"Organized analysis files into {analysis_dir}")
    return analysis_dir
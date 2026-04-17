import pandas as pd
import os
import glob
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

def merge_csv_files(directory_path):
    """合并目录下所有CSV文件并去除重复项"""
    print(f"正在处理目录: {directory_path}")
    
    # 获取目录下所有CSV文件
    csv_files = glob.glob(os.path.join(directory_path, "*.csv"))
    
    if not csv_files:
        print("没有找到CSV文件")
        return None
    
    print(f"找到 {len(csv_files)} 个CSV文件")
    
    # 分批读取CSV文件以减轻内存压力
    batch_size = 20
    total_batches = (len(csv_files) + batch_size - 1) // batch_size
    all_dfs = []
    total_records = 0
    
    for batch_idx in range(total_batches):
        start_idx = batch_idx * batch_size
        end_idx = min(start_idx + batch_size, len(csv_files))
        batch_files = csv_files[start_idx:end_idx]
        
        print(f"处理批次 {batch_idx+1}/{total_batches} (文件 {start_idx+1}-{end_idx}/{len(csv_files)})")
        
        # 读取批次中的所有文件
        batch_dfs = []
        for i, file in enumerate(batch_files):
            try:
                df = pd.read_csv(file, encoding='utf-8-sig')
                batch_dfs.append(df)
                total_records += len(df)
                
                # 每10个文件输出一次进度信息
                if (i + 1) % 10 == 0 or i == len(batch_files) - 1:
                    print(f"- 已读取 {i+1}/{len(batch_files)} 个文件")
            except Exception as e:
                print(f"读取文件 {file} 时出错: {e}")
        
        if batch_dfs:
            # 合并批次并立即去重
            batch_combined = pd.concat(batch_dfs, ignore_index=True)
            batch_combined = batch_combined.drop_duplicates().reset_index(drop=True)
            all_dfs.append(batch_combined)
            
            # 释放内存
            del batch_dfs
            del batch_combined
    
    if not all_dfs:
        print("没有成功读取任何文件")
        return None
    
    print(f"所有批次处理完成，合并最终结果...")
    
    # 合并所有批次结果
    combined_df = pd.concat(all_dfs, ignore_index=True)
    print(f"原始记录总数: {total_records}")
    print(f"合并后总记录数: {len(combined_df)}")
    
    # 最终去除重复项
    original_count = len(combined_df)
    combined_df = combined_df.drop_duplicates().reset_index(drop=True)
    print(f"去除重复后记录数: {len(combined_df)}, 移除了 {original_count - len(combined_df)} 条重复记录")
    
    # 移除空评论行
    original_count = len(combined_df)
    combined_df = combined_df.dropna(subset=['review']).reset_index(drop=True)
    print(f"移除空评论后记录数: {len(combined_df)}, 移除了 {original_count - len(combined_df)} 条空评论记录")
    
    return combined_df

def format_excel_file(df, output_path):
    """格式化并保存为Excel文件"""
    # 首字母大写列名
    df.columns = [col.capitalize() for col in df.columns]
    
    print("正在保存数据到Excel文件...")
    
    # 保存为Excel
    df.to_excel(output_path, index=False)
    print(f"保存Excel文件到: {output_path}")
    
    print("应用Excel格式...")
    
    # 应用格式
    wb = openpyxl.load_workbook(output_path)
    ws = wb.active
    
    # 设置标题行为粗体
    for cell in ws[1]:
        cell.font = Font(bold=True)
    
    # 设置列宽
    for i in range(1, 5):
        ws.column_dimensions[get_column_letter(i)].width = 15
    ws.column_dimensions['E'].width = 150
    
    # 设置单元格对齐方式和行高
    row_count = ws.max_row
    for row_idx, row in enumerate(ws.rows, 1):
        ws.row_dimensions[row[0].row].height = 35
        for i, cell in enumerate(row, 1):
            if i <= 4:
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                cell.alignment = Alignment(vertical='center')
        
        # 每5000行输出一次进度
        if row_idx % 5000 == 0 or row_idx == row_count:
            print(f"格式化进度: {row_idx}/{row_count} 行")
    
    # 保存格式化后的Excel
    wb.save(output_path)
    print(f"格式化Excel文件已保存: {output_path}")
    
    # 同时保存为txt文件
    txt_path = output_path.replace('.xlsx', '.txt')
    print(f"正在导出文本文件...")
    with open(txt_path, 'w', encoding='utf-8') as txt_file:
        txt_file.write('\t'.join(df.columns) + '\n')
        for idx, row in enumerate(df.iterrows(), 1):
            _, row_data = row
            row_values = [str(val) if pd.notna(val) else '' for val in row_data.values]
            txt_file.write('\t'.join(row_values) + '\n')
            
            # 每10000行输出一次进度
            if idx % 10000 == 0 or idx == len(df):
                print(f"文本导出进度: {idx}/{len(df)} 行")
                
    print(f"导出文本文件到: {txt_path}")
    
    return output_path

def main():
    # 输入目录路径
    directory_path = r"C:\Users\Alex\Downloads\oduho_com_trustpilot_data_20250323_194416\pages"
    
    # 输出文件路径
    output_dir = os.path.dirname(directory_path)
    brand_name = os.path.basename(output_dir).split('_trustpilot_data_')[0]
    output_excel = os.path.join(output_dir, f"{brand_name}_trustpilot_reviews_all.xlsx")
    
    # 合并CSV文件
    print("=" * 50)
    print("开始合并CSV文件...")
    print("=" * 50)
    combined_df = merge_csv_files(directory_path)
    
    if combined_df is not None and not combined_df.empty:
        # 格式化并保存Excel
        print("=" * 50)
        print("开始格式化并保存Excel文件...")
        print("=" * 50)
        format_excel_file(combined_df, output_excel)
        print("=" * 50)
        print("处理完成！")
        print(f"最终文件保存在: {output_excel}")
        print("=" * 50)
    else:
        print("处理失败，无法合并文件或结果为空")

if __name__ == "__main__":
    main() 
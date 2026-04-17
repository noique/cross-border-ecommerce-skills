import os
import pandas as pd
import glob
import datetime
import shutil
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

def merge_thread_data(base_dir):
    """合并所有线程目录下的CSV文件"""
    print(f"正在查找目录: {base_dir}")
    if not os.path.exists(base_dir):
        print(f"错误: 目录不存在 {base_dir}")
        return None

    # 寻找所有线程子目录
    thread_dirs = []
    for item in os.listdir(base_dir):
        thread_path = os.path.join(base_dir, item)
        if os.path.isdir(thread_path) and (
            item.startswith("local_ip") or 
            item.startswith("proxy_") or 
            item.startswith("v2ray_")
        ):
            thread_dirs.append(thread_path)
    
    if not thread_dirs:
        print("没有找到任何线程目录")
        return None
    
    print(f"找到 {len(thread_dirs)} 个线程目录:")
    for dir in thread_dirs:
        print(f"  - {os.path.basename(dir)}")
    
    # 查找每个线程目录下的pages子目录
    all_csv_files = []
    for thread_dir in thread_dirs:
        pages_dir = os.path.join(thread_dir, "pages")
        if os.path.exists(pages_dir):
            csv_files = glob.glob(os.path.join(pages_dir, "*.csv"))
            if csv_files:
                all_csv_files.extend(csv_files)
                print(f"在 {os.path.basename(thread_dir)}/pages 中找到 {len(csv_files)} 个CSV文件")
        else:
            print(f"警告: 没有找到pages目录: {pages_dir}")
    
    if not all_csv_files:
        print("没有找到任何CSV文件")
        return None
    
    print(f"总共找到 {len(all_csv_files)} 个CSV文件")
    
    # 读取所有CSV文件并合并
    all_dfs = []
    total_rows = 0
    
    for file in all_csv_files:
        try:
            df = pd.read_csv(file, encoding='utf-8-sig')
            all_dfs.append(df)
            total_rows += len(df)
            print(f"已读取: {os.path.basename(file)}, {len(df)} 行")
        except Exception as e:
            print(f"读取 {file} 时出错: {e}")
    
    if not all_dfs:
        print("没有成功读取任何文件")
        return None
    
    # 合并所有DataFrame
    merged_df = pd.concat(all_dfs, ignore_index=True)
    print(f"合并前总行数: {total_rows}")
    print(f"合并后总行数: {len(merged_df)}")
    
    # 去重
    original_count = len(merged_df)
    merged_df = merged_df.drop_duplicates(subset=['username', 'date', 'review']).reset_index(drop=True)
    print(f"去重后总行数: {len(merged_df)}, 删除了 {original_count - len(merged_df)} 条重复记录")
    
    # 移除空评论
    original_count = len(merged_df)
    merged_df = merged_df.dropna(subset=['review']).reset_index(drop=True)
    print(f"移除空评论后总行数: {len(merged_df)}, 删除了 {original_count - len(merged_df)} 条空评论")
    
    return merged_df

def format_excel_file(df, output_path):
    """格式化并保存为Excel文件"""
    # 首字母大写列名
    df.columns = [col.capitalize() for col in df.columns]
    
    print(f"正在保存数据到Excel文件: {output_path}")
    
    # 保存为Excel
    df.to_excel(output_path, index=False)
    
    print("应用Excel格式...")
    
    # 应用格式
    import openpyxl
    wb = openpyxl.load_workbook(output_path)
    ws = wb.active
    
    # 设置标题行为粗体
    for cell in ws[1]:
        cell.font = Font(bold=True)
    
    # 设置列宽
    for i in range(1, 6):  # 前5列
        if i == 5:  # 评论列
            ws.column_dimensions[get_column_letter(i)].width = 150
        else:
            ws.column_dimensions[get_column_letter(i)].width = 15
    
    # 设置单元格对齐方式和行高
    for row in ws.rows:
        ws.row_dimensions[row[0].row].height = 35
        for i, cell in enumerate(row, 1):
            if i < 5:  # 前4列居中
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:  # 评论列左对齐
                cell.alignment = Alignment(vertical='center')
    
    # 保存格式化后的Excel
    wb.save(output_path)
    print(f"格式化Excel文件已保存: {output_path}")
    
    # 同时保存为CSV和TXT文件
    csv_path = output_path.replace('.xlsx', '.csv')
    df.to_csv(csv_path, index=False, encoding='utf-8-sig')
    print(f"保存CSV文件: {csv_path}")
    
    txt_path = output_path.replace('.xlsx', '.txt')
    with open(txt_path, 'w', encoding='utf-8') as txt_file:
        txt_file.write('\t'.join(df.columns) + '\n')
        for _, row in df.iterrows():
            row_values = [str(val) if pd.notna(val) else '' for val in row.values]
            txt_file.write('\t'.join(row_values) + '\n')
    print(f"保存TXT文件: {txt_path}")
    
    return output_path

def copy_to_downloads(file_path):
    """将文件复制到用户的下载目录"""
    downloads_dir = os.path.join(os.path.expanduser("~"), "Downloads")
    if os.path.exists(downloads_dir):
        try:
            filename = os.path.basename(file_path)
            dest_path = os.path.join(downloads_dir, filename)
            shutil.copy2(file_path, dest_path)
            print(f"文件已复制到下载目录: {dest_path}")
            return dest_path
        except Exception as e:
            print(f"复制文件到下载目录时出错: {e}")
    return None

def main():
    # 查找最新的输出目录
    downloads_dir = os.path.join(os.path.expanduser("~"), "Downloads")
    trustpilot_dirs = [os.path.join(downloads_dir, d) for d in os.listdir(downloads_dir) 
                      if d.endswith("_trustpilot_data_" + datetime.datetime.now().strftime("%Y%m%d") + "*")]
    
    if not trustpilot_dirs:
        print("未找到今天的Trustpilot数据目录")
        # 询问用户输入目录路径
        user_dir = input("请输入Trustpilot数据目录的完整路径: ")
        if os.path.exists(user_dir):
            trustpilot_dirs = [user_dir]
        else:
            print(f"错误: 目录不存在 {user_dir}")
            return
    
    # 按修改时间排序，选择最新的目录
    trustpilot_dirs.sort(key=lambda x: os.path.getmtime(x), reverse=True)
    latest_dir = trustpilot_dirs[0]
    
    # 提取品牌名称
    brand_name = os.path.basename(latest_dir).split("_trustpilot_data_")[0]
    
    print(f"使用最新的数据目录: {latest_dir}")
    print(f"识别到的品牌名称: {brand_name}")
    
    # 合并所有线程数据
    merged_df = merge_thread_data(latest_dir)
    
    if merged_df is not None and not merged_df.empty:
        # 创建合并后的输出文件名
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        output_excel = os.path.join(latest_dir, f"{brand_name}_all_reviews_merged_{timestamp}.xlsx")
        
        # 格式化并保存Excel
        formatted_path = format_excel_file(merged_df, output_excel)
        
        # 复制到下载目录
        copy_to_downloads(formatted_path)
        
        print("=" * 50)
        print(f"处理完成! 共合并 {len(merged_df)} 条评论数据")
        print(f"最终文件保存在: {formatted_path}")
        print("=" * 50)
    else:
        print("处理失败，无法合并文件或结果为空")

if __name__ == "__main__":
    main() 
#!/usr/bin/env python3
"""Extract KOL/influencer prospects from Semrush backlink & refdomains xlsx files.

Usage:
    python3 extract_kol.py <input_dir> [--output <output.csv>] [--min-sources 2] [--min-traffic 500]

Arguments:
    input_dir       Directory containing Semrush xlsx files (can have subdirectories per domain)
    --output        Output CSV path (default: kol_prospects.csv in input_dir)
    --min-sources   Minimum number of competitor sources to include (default: 2)
    --min-traffic   Minimum organic traffic to include (default: 500)

Input files recognized:
    *-backlinks.xlsx           Main backlinks
    *-backlinks_refdomains.xlsx  Referring domains
    *-organic.Competitors-*.xlsx  Organic competitors
"""

import openpyxl
import os
import re
import csv
import sys
import argparse
from collections import defaultdict

# ── Skip lists ──

SKIP_PLATFORMS = {
    'google.com', 'facebook.com', 'youtube.com', 'twitter.com', 'instagram.com',
    'reddit.com', 'pinterest.com', 'tiktok.com', 'linkedin.com', 'wikipedia.org',
    'amazon.com', 'ebay.com', 'walmart.com', 'etsy.com', 'aliexpress.com',
    'apple.com', 'microsoft.com', 'yahoo.com', 'bing.com',
    'medium.com', 'substack.com', 'quora.com', 'stackoverflow.com',
    'imdb.com', 'indeed.com', 'tripadvisor.com', 'yelp.com', 'trustpilot.com',
    'shopify.com', 'wordpress.com', 'squarespace.com', 'wix.com',
    'nytimes.com', 'cnn.com', 'bbc.com', 'forbes.com',
    'x.com', 'fandom.com', 'mapquest.com', 'dictionary.com',
}

SKIP_BRAND_WORDS = {'google', 'facebook', 'youtube', 'twitter', 'instagram', 'reddit',
                    'pinterest', 'tiktok', 'linkedin', 'wikipedia', 'amazon', 'ebay',
                    'walmart', 'apple', 'microsoft', 'yahoo', 'bing'}

JUNK_PATTERNS = re.compile(r'blogspot\.com|wordpress\.com|\.info$|\.xyz$|\.tk$|\.ml$|\.ga$|\.cf$|\.gq$', re.I)

# ── Category patterns ──

CATEGORIES = {
    'KOL/Review': re.compile(r'review|blog|best|guide|test|unbox|everyday|daily|life|personal|vlog|howto|tip', re.I),
    'Media/Press': re.compile(r'news|press|magazine|journal|digest|wire|report|times|post|gazette|herald|tribune|media', re.I),
    'Forum/Community': re.compile(r'forum|talk|community|board|club|society|group|nation|hide|addicts|discussion', re.I),
    'Affiliate/Deal': re.compile(r'deals?|coupon|discount|compare|versus|vs|pick|choice|finder|shop|affiliate|partner', re.I),
}

# ── Social media & link-in-bio extractors ──

SOCIAL_PATTERNS = {
    # Link-in-bio services (richest source — one link → all socials)
    'linktree': re.compile(r'linktr\.ee/([a-zA-Z0-9_.]+)'),
    'linkbio': re.compile(r'(?:bio\.link|beacons\.ai|campsite\.bio|linkpop\.com|solo\.to|tap\.bio|lnk\.bio|milkshake\.app|carrd\.co|stan\.store|hoo\.be|withkoji\.com|snipfeed\.co)/([a-zA-Z0-9_.]+)'),
    # Social platforms
    'youtube': re.compile(r'youtube\.com/(?:c/|channel/|@|user/)([a-zA-Z0-9_.-]+)'),
    'instagram': re.compile(r'instagram\.com/([a-zA-Z0-9_.]+)'),
    'tiktok': re.compile(r'tiktok\.com/@([a-zA-Z0-9_.]+)'),
    'twitter': re.compile(r'(?:twitter\.com|x\.com)/([a-zA-Z0-9_]+)'),
    'facebook': re.compile(r'facebook\.com/(?:pg/|pages/)?([a-zA-Z0-9_.]+)'),
    'pinterest': re.compile(r'pinterest\.com/([a-zA-Z0-9_.]+)'),
    'twitch': re.compile(r'twitch\.tv/([a-zA-Z0-9_]+)'),
    'podcast': re.compile(r'(?:podcasts\.apple\.com|open\.spotify\.com/show|anchor\.fm)/([a-zA-Z0-9_-]+)'),
    'substack': re.compile(r'([a-zA-Z0-9_-]+)\.substack\.com'),
}

IG_SKIP = {'p', 'reel', 'stories', 'explore', 'accounts', 'about', 'directory', 'developer', 'legal'}
TWITTER_SKIP = {'intent', 'search', 'hashtag', 'i', 'share', 'home', 'explore', 'settings', 'login'}
FB_SKIP = {'sharer', 'share', 'dialog', 'login', 'groups', 'events', 'marketplace', 'watch', 'gaming'}

# ── Affiliate network patterns ──

AFFILIATE_PATTERNS = {
    'shareasale': re.compile(r'shareasale\.com', re.I),
    'impact': re.compile(r'impact\.com|impactradius\.com', re.I),
    'cj': re.compile(r'cj\.com|commission-junction', re.I),
    'awin': re.compile(r'awin\.com|awin1\.com', re.I),
    'rakuten': re.compile(r'rakutenadvertising\.com|linksynergy\.com', re.I),
    'partnerize': re.compile(r'partnerize\.com|prf\.hn', re.I),
    'avantlink': re.compile(r'avantlink\.com', re.I),
    'refersion': re.compile(r'refersion\.com', re.I),
    'goaffpro': re.compile(r'goaffpro\.com', re.I),
}


def should_skip(domain):
    if domain in SKIP_PLATFORMS:
        return True
    brand = domain.split('.')[0]
    if brand in SKIP_BRAND_WORDS:
        return True
    if domain.endswith('.gov') or domain.endswith('.edu') or domain.endswith('.mil'):
        return True
    if JUNK_PATTERNS.search(domain):
        return True
    return False


def categorize(domain):
    for cat, pattern in CATEGORIES.items():
        if pattern.search(domain):
            return cat
    return ''


def find_xlsx_files(input_dir):
    """Walk directory and find all relevant xlsx files."""
    files = {'backlinks': [], 'refdomains': [], 'competitors': []}
    for root, dirs, filenames in os.walk(input_dir):
        for f in filenames:
            if not f.endswith('.xlsx'):
                continue
            fl = f.lower()
            fpath = os.path.join(root, f)
            # Detect domain from parent folder
            parent = os.path.basename(root)
            if 'refdomains' in fl:
                files['refdomains'].append((fpath, parent))
            elif fl.endswith('-backlinks.xlsx') and 'anchor' not in fl and 'page' not in fl:
                files['backlinks'].append((fpath, parent))
            elif 'competitors' in fl:
                files['competitors'].append((fpath, parent))
    return files


def process_refdomains(files):
    """Extract domains from refdomains files."""
    domains = defaultdict(lambda: {'backlinks': 0, 'sources': set(), 'category': ''})

    for fpath, source_domain in files:
        try:
            wb = openpyxl.load_workbook(fpath, read_only=True)
            ws = wb.active
            headers = None
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    headers = [str(h).strip() if h else '' for h in row]
                    continue
                if not headers:
                    continue
                data = dict(zip(headers, row))
                domain = str(data.get('Root Domain / Category', data.get('Domain', '')) or '').strip().lower()
                bl = int(data.get('Backlinks', 0) or 0)
                if not domain or len(domain) < 4 or should_skip(domain):
                    continue
                entry = domains[domain]
                entry['backlinks'] += bl
                entry['sources'].add(source_domain)
                if not entry['category']:
                    entry['category'] = categorize(domain)
            wb.close()
        except Exception as e:
            print(f"  Warning: Error reading {fpath}: {e}", file=sys.stderr)

    return domains


def process_backlinks(files):
    """Extract social media handles, affiliate networks, and review sites from backlinks."""
    social = defaultdict(set)  # {platform: set of usernames}
    affiliates = defaultdict(lambda: {'urls': set(), 'sources': set()})  # affiliate network mentions
    review_domains = defaultdict(lambda: {'urls': [], 'sources': set()})

    for fpath, source_domain in files:
        try:
            wb = openpyxl.load_workbook(fpath, read_only=True)
            ws = wb.active
            headers = None
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    headers = [str(h).strip() if h else '' for h in row]
                    continue
                if not headers:
                    continue
                data = dict(zip(headers, row))
                url = str(data.get('Source url', data.get('Source URL', '')) or '')
                anchor = str(data.get('Anchor', '') or '').lower()

                # Extract social handles from URLs
                for platform, pattern in SOCIAL_PATTERNS.items():
                    match = pattern.search(url)
                    if match:
                        username = match.group(1)
                        if platform == 'instagram' and username in IG_SKIP:
                            continue
                        if platform == 'twitter' and username in TWITTER_SKIP:
                            continue
                        if platform == 'facebook' and username in FB_SKIP:
                            continue
                        social[platform].add(username)

                # Also check anchor text for social handles (e.g., "@username")
                at_match = re.findall(r'@([a-zA-Z0-9_]{3,30})', anchor)
                for handle in at_match:
                    if handle.lower() not in {'gmail', 'yahoo', 'hotmail', 'outlook', 'example'}:
                        social['mentioned_handle'].add(handle)

                # Detect affiliate network URLs (shows which brands use affiliate programs)
                for network, pattern in AFFILIATE_PATTERNS.items():
                    if pattern.search(url):
                        domain_match = re.search(r'https?://([^/]+)', url)
                        source_site = domain_match.group(1).replace('www.', '') if domain_match else url[:60]
                        affiliates[network]['urls'].add(source_site)
                        affiliates[network]['sources'].add(source_domain)

                # Extract review/blog/partner page URLs
                if re.search(r'review|best|top-?\d|guide|recommend|vs|comparison|unbox|partner|sponsor|collab|affiliate|ambassador', url, re.I):
                    domain_match = re.search(r'https?://([^/]+)', url)
                    if domain_match:
                        d = domain_match.group(1).replace('www.', '').lower()
                        if '.' in d and len(d) > 5 and not JUNK_PATTERNS.search(d) and not should_skip(d):
                            review_domains[d]['urls'].append(url[:120])
                            review_domains[d]['sources'].add(source_domain)
            wb.close()
        except Exception as e:
            print(f"  Warning: Error reading {fpath}: {e}", file=sys.stderr)

    return social, review_domains, affiliates


def process_competitors(files):
    """Extract competitor domains and traffic data."""
    domains = defaultdict(lambda: {'traffic': 0, 'common_kw': 0, 'relevance': 0, 'sources': set()})

    for fpath, source_domain in files:
        try:
            wb = openpyxl.load_workbook(fpath, read_only=True)
            ws = wb.active
            headers = None
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    headers = [str(h).strip() if h else '' for h in row]
                    continue
                if not headers:
                    continue
                data = dict(zip(headers, row))
                domain = str(data.get('Domain', '') or '').strip().lower()
                if not domain or should_skip(domain):
                    continue
                entry = domains[domain]
                entry['traffic'] = max(entry['traffic'], int(data.get('Organic Traffic', 0) or 0))
                entry['common_kw'] += int(data.get('Common Keywords', 0) or 0)
                entry['relevance'] = max(entry['relevance'], float(data.get('Competitor Relevance', 0) or 0))
                entry['sources'].add(source_domain)
            wb.close()
        except Exception as e:
            print(f"  Warning: Error reading {fpath}: {e}", file=sys.stderr)

    return domains


def main():
    parser = argparse.ArgumentParser(description='Extract KOL/influencer prospects from Semrush data')
    parser.add_argument('input_dir', help='Directory with Semrush xlsx files')
    parser.add_argument('--output', '-o', help='Output CSV path')
    parser.add_argument('--min-sources', type=int, default=2, help='Min competitor sources (default: 2)')
    parser.add_argument('--min-traffic', type=int, default=500, help='Min organic traffic (default: 500)')
    args = parser.parse_args()

    if not os.path.isdir(args.input_dir):
        print(f"Error: {args.input_dir} is not a directory", file=sys.stderr)
        sys.exit(1)

    output = args.output or os.path.join(args.input_dir, 'kol_prospects.csv')

    print(f"Scanning: {args.input_dir}")
    files = find_xlsx_files(args.input_dir)
    print(f"  Found: {len(files['refdomains'])} refdomains, {len(files['backlinks'])} backlinks, {len(files['competitors'])} competitors")

    # Process all data sources
    print("Processing refdomains...")
    ref_domains = process_refdomains(files['refdomains'])

    print("Processing backlinks...")
    social, review_domains, affiliates = process_backlinks(files['backlinks'])

    print("Processing competitors...")
    comp_domains = process_competitors(files['competitors'])

    # Merge all data
    all_domains = defaultdict(lambda: {
        'traffic': 0, 'backlinks': 0, 'sources': set(),
        'category': '', 'source_type': '', 'priority': 0,
    })

    for domain, data in ref_domains.items():
        entry = all_domains[domain]
        entry['backlinks'] += data['backlinks']
        entry['sources'] |= data['sources']
        if data['category']:
            entry['category'] = data['category']
        entry['source_type'] = 'refdomains'

    for domain, data in comp_domains.items():
        entry = all_domains[domain]
        entry['traffic'] = max(entry['traffic'], data['traffic'])
        entry['sources'] |= data['sources']
        if not entry['category']:
            entry['category'] = categorize(domain)
        if not entry['source_type']:
            entry['source_type'] = 'competitors'

    for domain, data in review_domains.items():
        entry = all_domains[domain]
        entry['sources'] |= data['sources']
        if not entry['category']:
            entry['category'] = 'KOL/Review'

    # Filter and score
    results = []
    for domain, data in all_domains.items():
        if len(data['sources']) < args.min_sources:
            continue
        if not data['category']:
            data['category'] = categorize(domain)

        # Priority score: sources * 10 + log(traffic) + category bonus
        import math
        cat_bonus = {'KOL/Review': 20, 'Media/Press': 15, 'Affiliate/Deal': 10, 'Forum/Community': 5}
        priority = (len(data['sources']) * 10 +
                    (math.log10(data['traffic'] + 1) * 3 if data['traffic'] > 0 else 0) +
                    cat_bonus.get(data['category'], 0))

        results.append({
            'domain': domain,
            'category': data['category'] or 'Other',
            'priority_score': round(priority, 1),
            'traffic': data['traffic'],
            'backlinks': data['backlinks'],
            'source_count': len(data['sources']),
            'sources': ', '.join(sorted(data['sources'])[:5]),
        })

    results.sort(key=lambda x: -x['priority_score'])

    # Write CSV
    with open(output, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.DictWriter(f, fieldnames=['domain', 'category', 'priority_score', 'traffic', 'backlinks', 'source_count', 'sources'])
        writer.writeheader()
        writer.writerows(results)

    # Print summary
    print(f"\n{'='*70}")
    print(f"  Results: {len(results)} prospects → {output}")
    print(f"{'='*70}")

    by_cat = defaultdict(list)
    for r in results:
        by_cat[r['category']].append(r)

    for cat in ['KOL/Review', 'Media/Press', 'Forum/Community', 'Affiliate/Deal', 'Other']:
        items = by_cat.get(cat, [])
        if not items:
            continue
        print(f"\n  [{cat}] ({len(items)})")
        for r in items[:8]:
            print(f"    {r['domain']:40s} score={r['priority_score']:>5}  traffic={r['traffic']:>8,}  sources={r['source_count']}")

    # Print social handles
    if social:
        print(f"\n  Social Media Handles Found:")
        for platform, handles in sorted(social.items()):
            if handles:
                sample = ', '.join(sorted(handles)[:10])
                more = f" (+{len(handles)-10} more)" if len(handles) > 10 else ''
                print(f"    [{platform}] ({len(handles)}): {sample}{more}")

    # Print affiliate networks
    if affiliates:
        print(f"\n  Affiliate Networks Detected:")
        for network, data in sorted(affiliates.items()):
            if data['urls']:
                print(f"    [{network}] found in {len(data['sources'])} competitor backlinks")

    # Write social handles to separate CSV
    social_output = output.replace('.csv', '_social.csv')
    with open(social_output, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.writer(f)
        writer.writerow(['platform', 'username', 'profile_url'])
        for platform, handles in sorted(social.items()):
            for handle in sorted(handles):
                url_map = {
                    'linktree': f'https://linktr.ee/{handle}',
                    'linkbio': f'https://bio.link/{handle}',
                    'youtube': f'https://youtube.com/@{handle}',
                    'instagram': f'https://instagram.com/{handle}',
                    'tiktok': f'https://tiktok.com/@{handle}',
                    'twitter': f'https://x.com/{handle}',
                    'facebook': f'https://facebook.com/{handle}',
                    'pinterest': f'https://pinterest.com/{handle}',
                    'twitch': f'https://twitch.tv/{handle}',
                    'substack': f'https://{handle}.substack.com',
                }
                profile = url_map.get(platform, handle)
                writer.writerow([platform, handle, profile])

    print(f"\n  Social handles CSV: {social_output}")
    print(f"\nDone. Main CSV: {output}")


if __name__ == '__main__':
    main()
